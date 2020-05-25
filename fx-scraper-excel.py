import requests
from bs4 import BeautifulSoup
import time 
from openpyxl import *
import datefinder

# Function to get the price string
def priceTracker():
	url = 'https://finance.yahoo.com/quote/AUDUSD=X?p=AUDUSD=X&.tsrc=fin-srch'

	response = requests.get(url)
	soup = BeautifulSoup(response.text,'lxml')
	price = soup.find_all('span', {'class':'Trsdu(0.3s)'})[0].text	
	return float(price)

# Function to get the time string
def TimeTracker():
	url = 'https://finance.yahoo.com/quote/AUDUSD=X?p=AUDUSD=X&.tsrc=fin-srch'

	response = requests.get(url)
	soup = BeautifulSoup(response.text,'lxml')
	time = soup.find_all("div", class_="Mstart(10px)")[0].text	
	matches =  datefinder.find_dates(time)
	for match in matches:
		#print(match)   # Unresolved: Why does this function print 3 times??? 
		x = match
		return x


# Preparing the excel workbook
wb = load_workbook('testbook.xlsx')                               #wb = Workbook()  #code to create new workbook
ws1 = wb.active
ws1.title = "USD AUD Price"
ws1["A1"] = "USD/AUD price"
ws1["B1"] = "Time"
col = 0
row = 2


while row<100:	  #this makes it run continuously till the 10th row

	print("'Current USD/AUD price is ' {}".format(priceTracker()))
	print("{}".format(TimeTracker()))

	ws1.cell(row=row , column=1, value=priceTracker())
	ws1.cell(row=row , column=2, value=TimeTracker())
	row = row + 1 
	time.sleep(55)   #Sleep duration in seconds

wb.save('testbook.xlsx')




